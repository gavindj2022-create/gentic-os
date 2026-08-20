"""Caller Service — load the Ascendance broker queue and place outbound AI calls.

MVP scope:
- Reads the Ascendance call sheet, pre-skips franchise/locked-site brokerages,
  quarantines VERIFY/blank numbers (never auto-dialed).
- Places calls via Bella's Retell client (override_agent_id + dynamic variables).
- Streams status + transcript over the GenTIC WebSocket (poll loop; webhook also
  supported when a public tunnel is available).
- Simulation mode (no Retell key/agent) so the cockpit is fully demoable offline.
Auto-routing to calendar / follow-up / Discord is Phase 2 (manual dispositions now).
"""

from __future__ import annotations
import os

import asyncio
import json
import re
import uuid
from datetime import datetime
from pathlib import Path

try:
    from zoneinfo import ZoneInfo
    _CHICAGO = ZoneInfo("America/Chicago")
except Exception:  # pragma: no cover
    _CHICAGO = None

from ..ws_manager import ws_manager
from ..database import (
    create_voice_call, update_voice_call, get_voice_calls, get_voice_call,
)
from . import retell as retell_shim
from . import sheets_sync

HERE = Path(__file__).parent
CONFIG_PATH = HERE / "voice_config.json"
TEMPLATE_PATH = HERE / "salescaller_prompt_template.md"
CALL_SHEET = Path(
    os.getenv("VAULT_PATH", "./vault") + "/Projects/Outbound Campaign A/"
    "Ascendance Call Sheet.xlsx"
)
BELLA_SHEET = Path(
    os.getenv("VAULT_PATH", "./vault") + "/Projects/Outbound Campaign B/"
    "Bella Prospect Call Sheet.xlsx"
)

# Calling window (prospect local time, America/Chicago).
CALL_HOUR_START = 8
CALL_HOUR_END = 20

# Franchise / HQ-locked brands to pre-skip even if they slip through the fit rating.
FRANCHISE_SKIP = [
    "compass", "keller williams", "re/max", "remax", "baird & warner",
    "baird and warner", "@properties", "atproperties", "jameson",
    "sotheby", "related realty", "coldwell", "century 21", "berkshire",
]

# Per-agent fields (each entry in cfg["agents"]).
_AGENT_FIELDS = ("label", "agent_id", "llm_id", "voice_id", "from_number",
                 "caller_name", "campaign", "price", "lead_source", "offer_summary")

# Shared, campaign-agnostic settings + the switchable agent roster. Each agent is
# a full caller persona (Retell agent + voice + lead source + pitch). The cockpit
# picks `active_agent`; place_call uses it (or a per-call override).
DEFAULT_CONFIG = {
    "active_agent": "bella",
    "demo_owner": "Gav",
    "callback_number": "+13123131478",
    "kill_switch": False,
    "simulate": False,           # force simulation even if a key is present
    "dnc": [],                   # E.164 numbers that must never be dialed
    "google_sheet_id": "",
    "google_sheet_url": "",
    "sheet_webhook_url": "",
    "agents": {
        "bella": {
            "label": "Bella — sells herself",
            "agent_id": "agent_b466cde33c9d7f871eaa0b9a36",
            "llm_id": "llm_b55d0430ffb6f96fe5f6a80c9e29",
            "voice_id": "11labs-Marissa",
            "from_number": "+13123131478",
            "caller_name": "Bella",
            "campaign": "bella",
            "price": "$147 a month",
            "lead_source": "bella",
            "offer_summary": (
                "Bella is an AI receptionist that answers every call 24/7, sounds "
                "like a real person, books callers straight into the calendar, and "
                "texts back missed calls — free 2-week pilot, then $147/mo."
            ),
        },
        "ascendance": {
            "label": "Alex — Ascendance AI",
            "agent_id": "agent_7a6cc664e70c914c96d79d5959",
            "llm_id": "llm_ef1f919dc5347b6e5d3e91fe8e34",
            "voice_id": "11labs-Adrian",
            "from_number": "+13123131478",
            "caller_name": "Alex",
            "campaign": "ascendance",
            "price": "$750 to set up plus $150 to 200 a month",
            "lead_source": "ascendance",
            "offer_summary": (
                "Ascendance AI is a smart AI chat assistant that lives on a broker's "
                "website. It talks to visitors 24/7, captures their name, budget, "
                "timeline and contact info, logs the lead to a Google Sheet, emails "
                "the broker instantly, and books the appointment onto their calendar."
            ),
        },
    },
}


# ───────────────────────── config ─────────────────────────

def _deep_default() -> dict:
    return json.loads(json.dumps(DEFAULT_CONFIG))


def load_config() -> dict:
    cfg = _deep_default()
    if CONFIG_PATH.exists():
        try:
            stored = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            stored = {}
        if "agents" not in stored and stored.get("agent_id"):
            # Migrate an old flat single-agent config into agents.ascendance.
            asc = cfg["agents"]["ascendance"]
            for k in _AGENT_FIELDS:
                if k in stored:
                    asc[k] = stored[k]
            cfg["active_agent"] = "ascendance"
            for k in ("kill_switch", "simulate", "demo_owner", "callback_number",
                      "google_sheet_id", "google_sheet_url", "sheet_webhook_url", "dnc"):
                if k in stored:
                    cfg[k] = stored[k]
        else:
            stored_agents = stored.pop("agents", {})
            cfg.update(stored)
            for key, agent in stored_agents.items():
                base = dict(cfg["agents"].get(key, {}))
                base.update(agent)
                cfg["agents"][key] = base
    cfg.setdefault("dnc", [])
    return cfg


def save_config(updates: dict) -> dict:
    cfg = load_config()
    cfg.update({k: v for k, v in updates.items() if v is not None})
    CONFIG_PATH.write_text(json.dumps(cfg, indent=2), encoding="utf-8")
    return cfg


def get_active_agent(cfg: dict | None = None, key: str | None = None) -> dict:
    """Resolve a full agent persona, defaulting to the active one."""
    cfg = cfg or load_config()
    agents = cfg.get("agents", {})
    key = key or cfg.get("active_agent") or "bella"
    if key not in agents:
        key = next(iter(agents), "")
    agent = dict(agents.get(key, {}))
    agent["key"] = key
    agent.setdefault("demo_owner", cfg.get("demo_owner", "Gav"))
    agent.setdefault("callback_number", cfg.get("callback_number", ""))
    return agent


def list_agents() -> list[dict]:
    cfg = load_config()
    active = cfg.get("active_agent")
    out = []
    for key, a in cfg.get("agents", {}).items():
        out.append({
            "key": key,
            "label": a.get("label", key),
            "caller_name": a.get("caller_name", ""),
            "voice_id": a.get("voice_id", ""),
            "from_number": a.get("from_number", ""),
            "campaign": a.get("campaign", key),
            "price": a.get("price", ""),
            "lead_source": a.get("lead_source", key),
            "configured": bool(a.get("agent_id")),
            "active": key == active,
        })
    return out


def set_active_agent(key: str) -> dict:
    cfg = load_config()
    if key not in cfg.get("agents", {}):
        return {"error": f"Unknown agent '{key}'."}
    save_config({"active_agent": key})
    return {"active_agent": key, "agents": list_agents()}


def set_agent_voice(key: str, voice_id: str) -> dict:
    """Update an agent's voice both locally and on Retell."""
    cfg = load_config()
    agents = cfg.get("agents", {})
    if key not in agents:
        return {"error": f"Unknown agent '{key}'."}
    agents[key]["voice_id"] = voice_id
    save_config({"agents": agents})
    agent = agents[key]
    if agent.get("agent_id") and retell_shim.has_key():
        try:
            client = retell_shim.get_client(live=True)
            client.update_agent(agent["agent_id"], {"voice_id": voice_id})
        except Exception as exc:
            return {"ok": True, "voice_id": voice_id, "warning": f"saved locally; Retell update failed: {exc}"}
    return {"ok": True, "voice_id": voice_id}


# ── Do-Not-Call list ──

def is_dnc(phone: str, cfg: dict | None = None) -> bool:
    cfg = cfg or load_config()
    n = normalize_phone(phone or "")
    return bool(n and n in set(cfg.get("dnc", [])))


def add_dnc(phone: str) -> dict:
    n = normalize_phone(phone or "") or (phone or "").strip()
    cfg = load_config()
    dnc = list(cfg.get("dnc", []))
    if n and n not in dnc:
        dnc.append(n)
        save_config({"dnc": dnc})
    return {"dnc": dnc}


def remove_dnc(phone: str) -> dict:
    n = normalize_phone(phone or "") or (phone or "").strip()
    cfg = load_config()
    dnc = [x for x in cfg.get("dnc", []) if x != n]
    save_config({"dnc": dnc})
    return {"dnc": dnc}


def set_kill_switch(active: bool) -> dict:
    return save_config({"kill_switch": bool(active)})


async def sync_call_record(call_id: str) -> dict:
    call = get_voice_call(call_id)
    if not call:
        return {"ok": False, "reason": "call not found"}
    result = await sheets_sync.sync_call(call)
    await ws_manager.broadcast("voice_sheet_sync", {"call_id": call_id, **sheets_sync.status()})
    return result


async def sync_all_calls(limit: int = 200) -> dict:
    calls = list(reversed(get_voice_calls(limit)))
    results = []
    for call in calls:
        results.append(await sheets_sync.sync_call(call))
    flushed = await sheets_sync.flush_pending()
    await ws_manager.broadcast("caller_status", get_status())
    return {"calls": len(calls), "results": results, "flush": flushed, **sheets_sync.status()}


# ───────────────────────── leads ─────────────────────────

def _stars(value) -> int:
    return str(value or "").count("★")  # ★


def normalize_phone(raw: str) -> str | None:
    """Return E.164 (+1XXXXXXXXXX) or None if not a usable US number."""
    digits = re.sub(r"\D", "", raw or "")
    if len(digits) == 10:
        return f"+1{digits}"
    if len(digits) == 11 and digits.startswith("1"):
        return f"+{digits}"
    return None


def load_ascendance_leads() -> dict:
    """Parse the call sheet into callable / needs_verification / skipped buckets."""
    result = {"callable": [], "needs_verification": [], "skipped": [], "error": None}
    if not CALL_SHEET.exists():
        result["error"] = f"Call sheet not found: {CALL_SHEET}"
        return result
    try:
        import openpyxl
    except Exception as exc:
        result["error"] = f"openpyxl not installed: {exc}"
        return result

    wb = openpyxl.load_workbook(CALL_SHEET, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    # Find the header row (contains both "Brokerage" and "Phone").
    header_idx = None
    for i, row in enumerate(rows):
        cells = [str(c or "").strip().lower() for c in row]
        if "brokerage" in cells and "phone" in cells:
            header_idx = i
            break
    if header_idx is None:
        result["error"] = "Could not locate header row in call sheet"
        return result

    for row in rows[header_idx + 1:]:
        fit_cell, business, phone_raw, website = (
            row[0], row[1], row[2] if len(row) > 2 else "",
            row[3] if len(row) > 3 else "",
        )
        phone_status = row[4] if len(row) > 4 else ""
        best_time = row[5] if len(row) > 5 else ""
        business = str(business or "").strip()
        if not business:
            continue
        if business.upper().startswith("TALLY"):
            break

        fit = _stars(fit_cell)
        lead = {
            "lead_id": re.sub(r"[^a-z0-9]+", "-", business.lower()).strip("-"),
            "business": business,
            "lead_name": "",  # owner unknown from the sheet; agent asks via gatekeeper
            "city": "Chicago",
            "fit": fit,
            "website": str(website or "").strip(),
            "best_time": str(best_time or "").strip(),
            "phone_raw": str(phone_raw or "").strip(),
        }

        name_l = business.lower()
        is_franchise = fit == 1 or any(f in name_l for f in FRANCHISE_SKIP)
        normalized = normalize_phone(lead["phone_raw"])
        needs_verify = (
            "verify" in str(phone_status or "").lower()
            or "verify" in lead["phone_raw"].lower()
            or normalized is None
        )

        if is_franchise:
            lead["skip_reason"] = "franchise / HQ-locked site"
            result["skipped"].append(lead)
        elif needs_verify:
            lead["verify_reason"] = "number needs verification"
            result["needs_verification"].append(lead)
        else:
            lead["phone"] = normalized
            result["callable"].append(lead)

    # Best fit first.
    result["callable"].sort(key=lambda x: -x["fit"])
    return result


def load_bella_leads() -> dict:
    """Parse the Bella prospect sheet (local service businesses to pitch Bella)."""
    result = {"callable": [], "needs_verification": [], "skipped": [], "error": None}
    if not BELLA_SHEET.exists():
        result["error"] = f"Bella prospect sheet not found: {BELLA_SHEET}"
        return result
    try:
        import openpyxl
    except Exception as exc:
        result["error"] = f"openpyxl not installed: {exc}"
        return result

    wb = openpyxl.load_workbook(BELLA_SHEET, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    header_idx, cols = None, {}
    for i, row in enumerate(rows):
        cells = [str(c or "").strip().lower() for c in row]
        if "business" in cells and "phone" in cells:
            header_idx, cols = i, {name: idx for idx, name in enumerate(cells)}
            break
    if header_idx is None:
        result["error"] = "Could not find a header row with Business + Phone columns."
        return result

    cfg = load_config()
    dnc = set(cfg.get("dnc", []))

    def cell(row, *names):
        for name in names:
            idx = cols.get(name)
            if idx is not None and idx < len(row):
                return row[idx]
        return ""

    for row in rows[header_idx + 1:]:
        business = str(cell(row, "business") or "").strip()
        if not business:
            continue
        phone_raw = str(cell(row, "phone") or "").strip()
        normalized = normalize_phone(phone_raw)
        status = str(cell(row, "status") or "").lower()
        lead = {
            "lead_id": re.sub(r"[^a-z0-9]+", "-", business.lower()).strip("-"),
            "business": business,
            "lead_name": "",
            "vertical": str(cell(row, "vertical") or "business").strip() or "business",
            "city": str(cell(row, "city") or "").strip() or "your area",
            "hook": str(cell(row, "hook") or "").strip(),
            "best_time": str(cell(row, "best time", "best_time") or "").strip(),
            "phone_raw": phone_raw,
            "fit": 3,
        }
        if normalized and normalized in dnc:
            lead["skip_reason"] = "on Do-Not-Call list"
            result["skipped"].append(lead)
        elif normalized is None or "verify" in status or "verify" in phone_raw.lower():
            lead["verify_reason"] = "number needs verification"
            result["needs_verification"].append(lead)
        else:
            lead["phone"] = normalized
            result["callable"].append(lead)
    return result


def load_leads(cfg: dict | None = None) -> dict:
    """Load the lead queue for the active agent's lead source."""
    cfg = cfg or load_config()
    source = get_active_agent(cfg).get("lead_source", "ascendance")
    if source == "bella":
        return load_bella_leads()
    return load_ascendance_leads()


# ───────────────────────── calling ─────────────────────────

def within_calling_hours() -> tuple[bool, str]:
    now = datetime.now(_CHICAGO) if _CHICAGO else datetime.now()
    if CALL_HOUR_START <= now.hour < CALL_HOUR_END:
        return True, ""
    return False, (
        f"Outside calling hours ({CALL_HOUR_START}:00–{CALL_HOUR_END}:00 "
        f"America/Chicago). Local time is {now.strftime('%H:%M')}."
    )


def _dynamic_vars(lead: dict, agent: dict) -> dict:
    return {
        "lead_name": lead.get("lead_name") or "there",
        "business": lead.get("business", ""),
        "city": lead.get("city", "your area"),
        "vertical": lead.get("vertical", "business") or "business",
        "hook": (lead.get("hook", "")
                 or "you're exactly the kind of local business I'm built for"),
        "caller_name": agent.get("caller_name", "Alex"),
        "demo_owner": agent.get("demo_owner", "Gav"),
        "callback_number": agent.get("callback_number", ""),
        "price": agent.get("price", ""),
        "offer_summary": agent.get("offer_summary", ""),
    }


def _should_simulate(cfg: dict, agent: dict) -> bool:
    if cfg.get("simulate"):
        return True
    available, _ = retell_shim.retell_available()
    # Go live only when everything the chosen agent needs is present.
    return not (available and agent.get("agent_id") and agent.get("from_number")
                and retell_shim.has_key())


async def place_call(lead: dict, override_hours: bool = False,
                     agent_key: str | None = None) -> dict:
    """Place (or simulate) an outbound call to a single lead via the chosen agent."""
    cfg = load_config()
    if cfg.get("kill_switch"):
        return {"error": "Kill switch is ON — calling is disabled."}
    agent = get_active_agent(cfg, agent_key)

    # Always coerce to E.164 — a queue lead's `phone` is already +1…, but a
    # Quick-Dial number arrives as raw user input and must be normalized too,
    # or Retell rejects it with a 400 ("not a valid number … e.164 format").
    phone = normalize_phone(lead.get("phone") or lead.get("phone_raw", ""))
    if not phone:
        raw = lead.get("phone") or lead.get("phone_raw", "")
        return {"error": f"Not a valid US phone number: {raw!r}. "
                         f"Use a 10-digit number, e.g. (312) 555-0143."}
    lead["phone"] = phone

    if is_dnc(phone, cfg):
        return {"error": f"{phone} is on the Do-Not-Call list — call blocked."}

    ok, reason = within_calling_hours()
    if not ok and not override_hours:
        return {"error": reason}

    call_id = uuid.uuid4().hex[:12]
    simulate = _should_simulate(cfg, agent)
    rec = create_voice_call(
        call_id,
        lead_name=lead.get("lead_name") or lead.get("business", ""),
        business=lead.get("business", ""),
        phone=phone,
        campaign=agent.get("campaign", agent.get("key", "bella")),
        simulated=simulate,
    )
    await ws_manager.broadcast("voice_call_update", {**rec, "transcript": ""})
    asyncio.create_task(sync_call_record(call_id))

    if simulate:
        asyncio.create_task(_simulate_call(call_id, lead, agent))
        return {"call_id": call_id, "status": "dialing", "simulated": True,
                "agent": agent.get("key")}

    try:
        client = retell_shim.get_client(live=True)
        resp = client.create_phone_call({
            "from_number": agent["from_number"],
            "to_number": phone,
            "override_agent_id": agent["agent_id"],
            "retell_llm_dynamic_variables": _dynamic_vars(lead, agent),
            "metadata": {"call_id": call_id, "business": lead.get("business", ""),
                         "agent": agent.get("key", "")},
        })
    except Exception as exc:
        update_voice_call(call_id, status="error", transcript=f"[error] {exc}",
                          ended_at=datetime.utcnow().isoformat())
        await ws_manager.broadcast("voice_call_update",
                                   {"call_id": call_id, "status": "error",
                                    "transcript": f"[error] {exc}"})
        asyncio.create_task(sync_call_record(call_id))
        return {"call_id": call_id, "error": str(exc)}

    retell_call_id = resp.get("call_id")
    update_voice_call(call_id, retell_call_id=retell_call_id)
    asyncio.create_task(sync_call_record(call_id))
    asyncio.create_task(_poll_call(call_id, retell_call_id))
    return {"call_id": call_id, "status": "dialing", "retell_call_id": retell_call_id,
            "simulated": False, "agent": agent.get("key")}


_STATUS_MAP = {"registered": "dialing", "ongoing": "live",
               "ended": "ended", "error": "error"}

# call_ids the operator has hung up — the sim loop checks this to stop early.
_ENDED: set[str] = set()


async def end_call(call_id: str) -> dict:
    """Hang up a live (or simulated) call from the cockpit."""
    call = get_voice_call(call_id)
    if not call:
        return {"error": "Call not found."}
    if call.get("status") in ("ended", "error"):
        return {"call_id": call_id, "status": call["status"]}

    _ENDED.add(call_id)  # signal the simulation loop to stop
    retell_call_id = call.get("retell_call_id")
    stop_error = None
    if retell_call_id and not call.get("simulated"):
        try:
            client = retell_shim.get_client(live=True)
            await asyncio.to_thread(client.stop_call, retell_call_id)
        except Exception as exc:
            # Retell may 4xx if the call already ended its side — we still want
            # the cockpit to reflect "ended", so record the note and continue.
            stop_error = str(exc)

    update_voice_call(call_id, status="ended", ended_at=datetime.utcnow().isoformat())
    asyncio.create_task(sync_call_record(call_id))
    await ws_manager.broadcast("voice_call_update", {"call_id": call_id, "status": "ended"})
    result = {"call_id": call_id, "status": "ended"}
    if stop_error:
        result["warning"] = stop_error
    return result


async def save_notes(call_id: str, notes: str) -> dict:
    """Persist operator notes for a call and mirror them to the Sheet."""
    if not get_voice_call(call_id):
        return {"error": "Call not found."}
    update_voice_call(call_id, notes=notes)
    asyncio.create_task(sync_call_record(call_id))
    await ws_manager.broadcast("voice_call_update", {"call_id": call_id, "notes": notes})
    return {"ok": True, "call_id": call_id, "notes": notes}


async def _poll_call(call_id: str, retell_call_id: str | None):
    """Poll Retell for status + transcript until the call ends (max ~10 min)."""
    if not retell_call_id:
        return
    client = retell_shim.get_client(live=True)
    last_transcript = ""
    for _ in range(300):  # 300 * 2s ≈ 10 min — poll fast so the transcript streams
        await asyncio.sleep(2)
        try:
            data = await asyncio.to_thread(client.get_call, retell_call_id)
        except Exception:
            continue
        status = _STATUS_MAP.get(data.get("call_status", ""), "live")
        transcript = data.get("transcript") or last_transcript
        fields = {"status": status, "transcript": transcript}
        if data.get("recording_url"):
            fields["recording_url"] = data["recording_url"]
        if status in ("ended", "error"):
            fields["ended_at"] = datetime.utcnow().isoformat()
        update_voice_call(call_id, **fields)
        asyncio.create_task(sync_call_record(call_id))
        if transcript != last_transcript or status in ("ended", "error"):
            await ws_manager.broadcast("voice_call_update", {"call_id": call_id, **fields})
            last_transcript = transcript
        if status in ("ended", "error"):
            return


_SIM_SCRIPT = [
    ("agent", "Hi, is this the team at {business}? This is {caller_name}, an AI "
              "assistant calling on behalf of Ascendance AI. I know it's a cold "
              "call, so I'll be quick — got 30 seconds?"),
    ("user", "Uh, sure, what's this about?"),
    ("agent", "Most people who visit a broker's website leave without ever "
              "reaching out — no name, no number. That's a commission walking "
              "out the door. We put an AI assistant on your site that talks to "
              "those visitors and books them onto your calendar, 24/7."),
    ("user", "We already have a website though."),
    ("agent", "Perfect — keep it exactly as it is. This isn't a website, it sits "
              "on top of it and captures the visitors you're losing right now. "
              "You make ten, fifteen grand on a sale; this is {price}. One caught "
              "client pays for it for years."),
    ("user", "Okay, that's interesting."),
    ("agent", "Let's grab 10 minutes this week — {demo_owner} will show you "
              "exactly how it works on a site like yours. Does earlier or later "
              "in the week work better?"),
    ("user", "Later in the week could work."),
    ("agent", "Great — I'll have {demo_owner} send a calendar invite. What's the "
              "best email? [SIMULATION — no real call placed]"),
]


async def _simulate_call(call_id: str, lead: dict, agent: dict):
    """Stream a canned conversation so the cockpit works without a Retell key."""
    dyn = _dynamic_vars(lead, agent)
    lines: list[str] = []
    await asyncio.sleep(1.2)
    update_voice_call(call_id, status="live")
    asyncio.create_task(sync_call_record(call_id))
    await ws_manager.broadcast("voice_call_update",
                               {"call_id": call_id, "status": "live", "transcript": ""})
    for who, text in _SIM_SCRIPT:
        await asyncio.sleep(1.8)
        if call_id in _ENDED:
            _ENDED.discard(call_id)
            return  # operator hung up mid-simulation
        speaker = "Agent" if who == "agent" else "Prospect"
        lines.append(f"{speaker}: {text.format(**dyn)}")
        transcript = "\n".join(lines)
        update_voice_call(call_id, transcript=transcript)
        asyncio.create_task(sync_call_record(call_id))
        await ws_manager.broadcast("voice_call_update",
                                   {"call_id": call_id, "status": "live",
                                    "transcript": transcript})
    await asyncio.sleep(1.2)
    update_voice_call(call_id, status="ended", ended_at=datetime.utcnow().isoformat())
    asyncio.create_task(sync_call_record(call_id))
    await ws_manager.broadcast("voice_call_update",
                               {"call_id": call_id, "status": "ended",
                                "transcript": "\n".join(lines)})


def _render_transcript(call: dict) -> str:
    """Build readable transcript text from a Retell call object."""
    if call.get("transcript"):
        return call["transcript"]
    obj = call.get("transcript_with_tool_calls") or call.get("transcript_object")
    if isinstance(obj, list):
        lines = []
        for t in obj:
            role, content = t.get("role"), t.get("content")
            if content and role in ("agent", "user"):
                who = "Agent" if role == "agent" else "Prospect"
                lines.append(f"{who}: {content}")
        return "\n".join(lines)
    return ""


async def handle_webhook(event: dict):
    """Handle a Retell webhook event.

    Handles `transcript_updated` (streams many times per call → live transcript),
    plus `call_started` / `call_ended` / `call_analyzed`. Requires the agent's
    webhook_url to point at a public tunnel to this endpoint.
    """
    name = event.get("event") or event.get("event_type") or ""
    call = event.get("call") or event.get("data") or {}
    retell_call_id = call.get("call_id")
    meta = call.get("metadata") or {}
    call_id = meta.get("call_id")
    if not call_id and retell_call_id:
        # Match by stored retell_call_id.
        for c in get_voice_calls(200):
            if c.get("retell_call_id") == retell_call_id:
                call_id = c["call_id"]
                break
    if not call_id:
        return {"ok": False, "reason": "unknown call"}

    fields: dict = {}
    if name in ("call_started", "transcript_updated"):
        fields["status"] = "live"
    elif name in ("call_ended", "call_analyzed"):
        fields["status"] = "ended"
        fields["ended_at"] = datetime.utcnow().isoformat()
    tr = _render_transcript(call)
    if tr:
        fields["transcript"] = tr
    if call.get("recording_url"):
        fields["recording_url"] = call["recording_url"]
    if fields:
        update_voice_call(call_id, **fields)
        asyncio.create_task(sync_call_record(call_id))
        await ws_manager.broadcast("voice_call_update", {"call_id": call_id, **fields})
    return {"ok": True}


# ───────────────────────── status ─────────────────────────

def spend_summary() -> dict:
    """Estimate Retell spend from logged call durations.

    Retell bills per minute (voice infra + LLM + TTS + telephony). We don't have
    their exact per-call cost stored, so this is an estimate: live-call minutes ×
    a configurable blended rate (default ~$0.18/min, from Gav's invoice). Plus the
    flat monthly phone-number rental.
    """
    cfg = load_config()
    rate = float(cfg.get("retell_rate_per_min", 0.18))
    number_monthly = float(cfg.get("retell_number_monthly", 2.0))
    today = datetime.utcnow().date().isoformat()

    def minutes(call: dict) -> float:
        try:
            s = datetime.fromisoformat(call["created_at"])
            e = datetime.fromisoformat(call["ended_at"])
            # Cap at 60 min/call — a longer span is an orphaned call that was
            # swept to 'ended' long after it really hung up, not real talk time.
            return min(60.0, max(0.0, (e - s).total_seconds()) / 60.0)
        except Exception:
            return 0.0

    tot_min = today_min = 0.0
    tot_calls = today_calls = 0
    for call in get_voice_calls(2000):
        if call.get("simulated") or not call.get("ended_at"):
            continue
        m = minutes(call)
        tot_min += m
        tot_calls += 1
        if str(call.get("created_at", "")).startswith(today):
            today_min += m
            today_calls += 1
    return {
        "estimated": True,
        "rate_per_min": rate,
        "number_monthly": number_monthly,
        "total_calls": tot_calls,
        "total_minutes": round(tot_min, 1),
        "total_cost": round(tot_min * rate, 2),
        "today_calls": today_calls,
        "today_minutes": round(today_min, 1),
        "today_cost": round(today_min * rate, 2),
        "month_estimate": round(tot_min * rate + number_monthly, 2),
    }


def get_status() -> dict:
    cfg = load_config()
    agent = get_active_agent(cfg)
    available, err = retell_shim.retell_available()
    return {
        "active_agent": agent.get("key"),
        "agent_label": agent.get("label"),
        "caller_name": agent.get("caller_name"),
        "voice_id": agent.get("voice_id"),
        "campaign": agent.get("campaign", agent.get("key")),
        "price": agent.get("price", ""),
        "lead_source": agent.get("lead_source", agent.get("key")),
        "agents": list_agents(),
        "kill_switch": cfg.get("kill_switch", False),
        "from_number": agent.get("from_number", ""),
        "demo_owner": cfg.get("demo_owner", "Gav"),
        "agent_configured": bool(agent.get("agent_id")),
        "retell_available": available,
        "retell_error": err,
        "has_api_key": retell_shim.has_key(),
        "simulate": _should_simulate(cfg, agent),
        "calling_hours_ok": within_calling_hours()[0],
        "dnc_count": len(cfg.get("dnc", [])),
        "sheet_sync": sheets_sync.status(),
    }
